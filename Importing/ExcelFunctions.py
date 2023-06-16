#This class provides the methods to handle excel documents
#Author: Caleb Harris - UCCS OIT Services Professional
#Date Created: 4/11/2022

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import CLI
import datetime
    
def getDataFromExcel():
    """
    Function gets data from excel and stores in passed in data structure
    """

    #define workbook
    file_path = CLI.getFilePath()
    inventoryWorkbook = load_workbook(filename = file_path)
    
    #define sheet
    print("\nExcel sheets in this file: \n------------------------------")
    sheetNames = inventoryWorkbook.sheetnames
    for sheet in sheetNames:
        print(sheet)
    print("------------------------------\n")
    valid_sheetname = False
    sheetname = CLI.getSheetName()
    while not valid_sheetname:
        try:
            incomingDataSheet = inventoryWorkbook[sheetname]
            valid_sheetname = True
        except KeyError:
            print("Invalid Input!")
            sheetname = CLI.getSheetName()

    #find rows and cols of sheet including header line
    rows = incomingDataSheet.max_row
    cols = incomingDataSheet.max_column

    #store data in object then store object in data structure
    data = []
    for row in range(rows):
        data.append([])
        for col in range(cols):
            colLetter = get_column_letter(col+1)
            currentHeaderVal = incomingDataSheet[f"{colLetter}1"].value
            currentCellVal = incomingDataSheet[f"{colLetter}{row + 1}"].value

            if "quantity" in currentHeaderVal.lower() and row != 0:
                if currentCellVal != None and currentCellVal != '':
                    currentCellVal = int(currentCellVal)
            elif "date" in currentHeaderVal.lower() and currentCellVal != None and currentCellVal != '':
                currentCellVal = datetime.strptime(currentCellVal, "%m//%d//%Y")
            if currentCellVal == '':
                currentCellVal = None

            data[row].append(currentCellVal)

    #test that all data is present
    assert(len(data) == rows), "Rows read from Excel Sheet dont match rows in data structure!"
    assert(len(data[0]) == cols), "Cols read from Excel Sheet dont match cols in data structure!"

    return data
